import customtkinter as ctk
import pandas as pd
from tkinter import filedialog, messagebox
import threading
import re
from openpyxl import load_workbook # Movido para o topo

# Configuração visual personalizada
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue") 

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Conciliador de Matrículas - Sistema DE/PARA")
        self.geometry("800x600")
        self.configure(fg_color="#0C3B13") 

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)

        self.label_titulo = ctk.CTkLabel(self, text="CONCILIAÇÃO DE MATRÍCULAS", 
                                         font=ctk.CTkFont(size=22, weight="bold"),
                                         text_color="#FFFFFF")
        self.label_titulo.grid(row=0, column=0, padx=20, pady=25)

        # Frame DE
        self.frame_rh = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color="#488441")
        self.frame_rh.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        self.label_de = ctk.CTkLabel(self.frame_rh, text="DE (RH):", font=ctk.CTkFont(weight="bold"))
        self.label_de.pack(side="left", padx=10)
        self.path_rh = ctk.CTkEntry(self.frame_rh, placeholder_text="Selecione o arquivo de origem...", width=450)
        self.path_rh.pack(side="left", padx=10, pady=15)
        self.btn_rh = ctk.CTkButton(self.frame_rh, text="Buscar Arquivo", fg_color="#1A237E", command=self.sel_rh)
        self.btn_rh.pack(side="right", padx=10)

        # Frame PARA
        self.frame_base = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color="#D1D1D1")
        self.frame_base.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.label_para = ctk.CTkLabel(self.frame_base, text="PARA (Base):", font=ctk.CTkFont(weight="bold"))
        self.label_para.pack(side="left", padx=10)
        self.path_base = ctk.CTkEntry(self.frame_base, placeholder_text="Selecione o arquivo de destino...", width=450)
        self.path_base.pack(side="left", padx=10, pady=15)
        self.btn_base = ctk.CTkButton(self.frame_base, text="Buscar Arquivo", fg_color="#1A237E", command=self.sel_base)
        self.btn_base.pack(side="right", padx=10)

        self.btn_run = ctk.CTkButton(self, text="EXECUTAR CONCILIAÇÃO", 
                                     fg_color="#28a745", hover_color="#218838",
                                     font=ctk.CTkFont(size=16, weight="bold"), 
                                     height=45, command=self.start_thread)
        self.btn_run.grid(row=3, column=0, padx=20, pady=25)

        self.log_text = ctk.CTkTextbox(self, width=760, height=200, border_width=1)
        self.log_text.grid(row=4, column=0, padx=20, pady=10, sticky="nsew")

    def add_log(self, text):
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")

    def sel_rh(self):
        self.path_rh.delete(0, "end")
        self.path_rh.insert(0, filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))

    def sel_base(self):
        self.path_base.delete(0, "end")
        self.path_base.insert(0, filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")]))

    def start_thread(self):
        threading.Thread(target=self.processar_dados, daemon=True).start()

    def processar_dados(self):
        try:
            p_rh = self.path_rh.get()
            p_base = self.path_base.get()

            if not p_rh or not p_base:
                messagebox.showwarning("Atenção", "Selecione os arquivos!")
                return

            self.add_log("> Lendo planilhas...")
            df_rh = pd.read_excel(p_rh)
            df_base = pd.read_excel(p_base)

            # Posições: Coluna A (0) e Coluna B (1)
            col_antiga_rh = df_rh.columns[0]
            col_nova_rh = df_rh.columns[1]
            col_antiga_base = df_base.columns[0]

            self.add_log(f"> Ordenando pela Coluna A...")
            df_rh = df_rh.sort_values(by=col_antiga_rh).reset_index(drop=True)
            df_base = df_base.sort_values(by=col_antiga_base).reset_index(drop=True)

            self.add_log("> Padronizando CPFs...")
            def clean_cpf(x): return re.sub(r'\D', '', str(x))
            col_cpf_rh = [c for c in df_rh.columns if 'CPF' in str(c).upper()][0]
            col_cpf_base = [c for c in df_base.columns if 'CPF' in str(c).upper()][0]

            df_rh['CPF_LIMPO'] = df_rh[col_cpf_rh].apply(clean_cpf)
            df_base['CPF_LIMPO'] = df_base[col_cpf_base].apply(clean_cpf)

            self.add_log("> Mapeando vínculos...")
            df_rh['id_unico'] = df_rh['CPF_LIMPO'] + "_" + df_rh.groupby('CPF_LIMPO').cumcount().add(1).astype(str)
            df_base['id_unico'] = df_base['CPF_LIMPO'] + "_" + df_base.groupby('CPF_LIMPO').cumcount().add(1).astype(str)

            self.add_log("> Cruzando informações...")
            df_rh_mini = df_rh[['id_unico', col_nova_rh]]
            df_merged = pd.merge(df_base, df_rh_mini, on='id_unico', how='left')

            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if save_path:
                self.add_log("> Gravando dados na Coluna B e mantendo formatação...")
                wb = load_workbook(p_base)
                ws = wb.active

                novas_matriculas = df_merged[col_nova_rh].tolist()

                for i, valor in enumerate(novas_matriculas, start=2):
                    celula = ws.cell(row=i, column=2)
                    if pd.notna(valor):
                        try:
                            celula.value = int(valor)
                        except:
                            celula.value = valor
                    else:
                        celula.value = ""

                wb.save(save_path)
                self.add_log(f"✅ SUCESSO: Arquivo salvo em {save_path}")
                messagebox.showinfo("Sucesso", "Conciliação finalizada!")
        
        except Exception as e:
            self.add_log(f"❌ ERRO: {str(e)}")

if __name__ == "__main__":
    app = App()
    app.mainloop()